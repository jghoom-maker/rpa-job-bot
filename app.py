import streamlit as st
import os
import smtplib
import openpyxl
from openpyxl.styles import Font, Alignment
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime
import time

st.set_page_config(page_title="📧 RPA Job Email Bot", page_icon="🤖", layout="wide")

st.title("📧 RPA Job Email Application Bot")
st.markdown("Automate your job applications with personalized, dynamic emails!")

# Create tabs
tab1, tab2, tab3 = st.tabs(["Configuration", "Email Template", "About"])

with tab1:
    st.header("⚙️ Configuration")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📧 Email Settings")
        sender_email = st.text_input("Gmail Address:", value="your_email@gmail.com")
        sender_password = st.text_input("Gmail App Password:", type="password", value="16-character password")
        sender_name = st.text_input("Your Name:", value="Your Full Name")
    
    with col2:
        st.subheader("💼 Professional Details")
        your_skills = st.text_area("Your Skills:", value="Python, Django, REST APIs", height=100)
        years_experience = st.text_input("Years of Experience:", value="3+ years")
    
    st.divider()
    
    st.subheader("📄 File Upload")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**Upload Excel File**")
        st.markdown("Expected columns: Company Name | Email | Job Title | Resume File")
        excel_file = st.file_uploader("Choose Excel file", type="xlsx", key="excel")
    
    with col2:
        st.markdown("**Upload Resume PDFs**")
        st.markdown("Upload all your resume PDF files")
        resume_files = st.file_uploader("Choose resume files", type="pdf", accept_multiple_files=True, key="resumes")
    
    st.divider()
    
    if st.button("✅ Verify Setup", key="verify"):
        issues = []
        
        if sender_email == "your_email@gmail.com" or "@" not in sender_email:
            issues.append("❌ Gmail address not configured")
        if sender_password == "16-character password" or len(sender_password) < 10:
            issues.append("❌ Gmail app password not set")
        if not excel_file:
            issues.append("❌ Excel file not uploaded")
        if not resume_files:
            issues.append("❌ Resume files not uploaded")
        
        if issues:
            st.error("\n".join(issues))
        else:
            st.success("✅ All settings configured correctly!")

with tab2:
    st.header("📝 Email Template")
    
    st.markdown("""
    ### Available Variables
    Insert these variables into your email template. They will be automatically replaced with actual values:
    
    - **{company_name}** - The company name
    - **{job_title}** - The job title you're applying for
    - **{years_experience}** - Your years of experience
    - **{your_skills}** - Your professional skills
    - **{sender_name}** - Your name (signature)
    """)
    
    st.divider()
    
    # Create columns for template editing and variables
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.subheader("✍️ Write Your Email Template")
        
        # Default template
        default_template = """Dear Hiring Manager at {company_name},

I am writing to express my strong interest in the {job_title} position at your esteemed organization.

With {years_experience} of professional experience, I have developed strong expertise in {your_skills}. I am confident that my technical skills and problem-solving abilities make me an excellent fit for your team.

I am particularly impressed by {company_name}'s commitment to innovation and would be excited to contribute to your projects. My attached resume details my professional background and accomplishments.

I would welcome the opportunity to discuss how my skills and experience can contribute to {company_name}'s continued success. Please feel free to contact me at your convenience.

Thank you for considering my application. I look forward to hearing from you.

Best regards,
{sender_name}"""
        
        email_template = st.text_area(
            "Email Body Template:",
            value=default_template,
            height=350,
            key="template"
        )
        
        # Store template in session
        st.session_state.email_template = email_template
    
    with col2:
        st.subheader("🔧 Insert Variables")
        
        variables = [
            ("{company_name}", "Company Name"),
            ("{job_title}", "Job Title"),
            ("{years_experience}", "Experience"),
            ("{your_skills}", "Skills"),
            ("{sender_name}", "Your Name")
        ]
        
        for var_code, var_label in variables:
            if st.button(f"Insert {var_label}", key=f"btn_{var_label}"):
                # Insert variable at cursor (append to end for simplicity)
                st.session_state.email_template = email_template + f"\n{var_code}"
                st.rerun()
    
    st.divider()
    
    st.subheader("👁️ Live Preview")
    st.markdown("Here's how your email will look (with sample data):")
    
    # Sample data for preview
    sample_data = {
        "company_name": "Google",
        "job_title": "Senior Software Engineer",
        "years_experience": "5+ years",
        "your_skills": "Python, Cloud Computing, System Design",
        "sender_name": "John Smith"
    }
    
    # Create preview by replacing variables
    preview_text = email_template
    for key, value in sample_data.items():
        preview_text = preview_text.replace(f"{{{key}}}", value)
    
    st.info(preview_text)

with tab3:
    st.header("ℹ️ About This Tool")
    st.markdown("""
    ### 🤖 RPA Job Email Bot - Professional Edition
    
    Automate your job applications with personalized, dynamic emails!
    
    #### ✨ Key Features:
    - ✅ Send emails to multiple companies
    - ✅ Use different resumes for different job titles
    - ✅ **Custom email templates with dynamic variables**
    - ✅ Personalized email content per company
    - ✅ Automatic PDF attachment
    - ✅ Detailed logging and tracking
    - ✅ Live email preview
    
    #### 📝 How to Create Your Email:
    1. Go to "Email Template" tab
    2. Write your custom email message
    3. Insert variables like {company_name}, {job_title} etc.
    4. Preview how it looks with sample data
    5. Variables are automatically replaced when sending
    
    #### 🚀 How to Use:
    1. Enter your Gmail credentials
    2. Fill in your professional details
    3. Create your custom email template
    4. Upload your companies.xlsx file
    5. Upload all your resume PDF files
    6. Click "Start Sending Applications"
    
    #### 📋 Setup Requirements:
    - Gmail account with 2-Step Verification enabled
    - Gmail App Password (16 characters)
    - Excel file with 4 columns:
      - Column A: Company Name
      - Column B: Email Address
      - Column C: Job Title
      - Column D: Resume File (PDF filename)
    - PDF resume files (named exactly as in Column D)
    
    #### 🔐 Gmail App Password Setup:
    1. Go to: https://myaccount.google.com/security
    2. Enable 2-Step Verification if not already enabled
    3. Click "App passwords" (under 2-Step Verification)
    4. Select: Mail + Windows Computer (or your device)
    5. Google generates a 16-character password
    6. Copy and paste into the app
    
    #### 💡 Email Template Tips:
    - Use {company_name} to personalize for each company
    - Use {job_title} to reference the specific position
    - Be professional but authentic
    - Keep it concise (3-4 paragraphs)
    - Always include a strong closing
    
    Made with ❤️ using Streamlit and Python
    """)

st.divider()

# Send Emails Section (at bottom)
st.header("🚀 Send Applications")

col1, col2 = st.columns(2)

with col1:
    if st.button("🚀 START SENDING APPLICATIONS", key="send", use_container_width=True):
        # Get values from session
        if "email_template" not in st.session_state:
            st.session_state.email_template = default_template
        
        email_template = st.session_state.email_template
        
        # Get configuration values (need to re-read from tab1)
        sender_email_val = st.session_state.get('sender_email', 'your_email@gmail.com')
        sender_password_val = st.session_state.get('sender_password', '')
        sender_name_val = st.session_state.get('sender_name', '')
        your_skills_val = st.session_state.get('your_skills', '')
        years_exp_val = st.session_state.get('years_experience', '')
        excel_file_val = st.session_state.get('excel_file', None)
        resume_files_val = st.session_state.get('resume_files', [])
        
        if sender_email_val == "your_email@gmail.com":
            st.error("❌ Please enter your Gmail address in Configuration tab")
        elif not sender_password_val or sender_password_val == "16-character password":
            st.error("❌ Please enter your Gmail app password in Configuration tab")
        elif not sender_name_val:
            st.error("❌ Please enter your name in Configuration tab")
        elif not excel_file_val:
            st.error("❌ Please upload your Excel file in Configuration tab")
        elif not resume_files_val:
            st.error("❌ Please upload resume files in Configuration tab")
        else:
            # Save uploaded files temporarily
            with open("companies.xlsx", "wb") as f:
                f.write(excel_file_val.getbuffer())
            
            resume_folder = "resumes"
            os.makedirs(resume_folder, exist_ok=True)
            
            for resume in resume_files_val:
                with open(os.path.join(resume_folder, resume.name), "wb") as f:
                    f.write(resume.getbuffer())
            
            st.success("✅ Files saved successfully!")
            
            # Read Excel
            try:
                wb = openpyxl.load_workbook("companies.xlsx")
                ws = wb.active
                
                companies = []
                for row in range(2, 1000):
                    company_name = ws[f'A{row}'].value
                    company_email = ws[f'B{row}'].value
                    job_title = ws[f'C{row}'].value
                    resume_file = ws[f'D{row}'].value
                    
                    if company_name is None or company_email is None:
                        break
                    
                    companies.append({
                        'name': company_name,
                        'email': company_email,
                        'job_title': job_title if job_title else "Position",
                        'resume_file': resume_file if resume_file else "resume.pdf"
                    })
                
                if not companies:
                    st.error("❌ No companies found in Excel file!")
                else:
                    st.info(f"📋 Found {len(companies)} companies")
                    
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    sent_count = 0
                    failed_count = 0
                    sent_log = []
                    
                    for i, company in enumerate(companies):
                        progress = (i + 1) / len(companies)
                        progress_bar.progress(progress)
                        
                        company_name = company['name']
                        company_email = company['email']
                        job_title = company['job_title']
                        resume_file = company['resume_file']
                        
                        status_text.text(f"[{i+1}/{len(companies)}] Sending to {company_name}...")
                        
                        # Check resume exists
                        resume_path = os.path.join(resume_folder, resume_file)
                        if not os.path.exists(resume_path):
                            status_text.text(f"[{i+1}/{len(companies)}] ⚠️ Resume not found: {resume_file}")
                            failed_count += 1
                            sent_log.append({
                                'company': company_name,
                                'job_title': job_title,
                                'resume': resume_file,
                                'status': '❌ Resume not found'
                            })
                            continue
                        
                        try:
                            # Create personalized email body
                            email_body = email_template
                            email_body = email_body.replace("{company_name}", company_name)
                            email_body = email_body.replace("{job_title}", job_title)
                            email_body = email_body.replace("{years_experience}", years_exp_val)
                            email_body = email_body.replace("{your_skills}", your_skills_val)
                            email_body = email_body.replace("{sender_name}", sender_name_val)
                            
                            # Create email
                            message = MIMEMultipart()
                            message['From'] = sender_name_val
                            message['To'] = company_email
                            message['Subject'] = f"Application for {job_title} Position"
                            message.attach(MIMEText(email_body, 'plain'))
                            
                            # Attach resume
                            with open(resume_path, 'rb') as attachment:
                                part = MIMEBase('application', 'octet-stream')
                                part.set_payload(attachment.read())
                            
                            encoders.encode_base64(part)
                            part.add_header('Content-Disposition', f'attachment; filename= {resume_file}')
                            message.attach(part)
                            
                            # Send email
                            server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
                            server.login(sender_email_val, sender_password_val)
                            server.send_message(message)
                            server.quit()
                            
                            status_text.text(f"[{i+1}/{len(companies)}] ✅ Sent to {company_name}")
                            sent_count += 1
                            sent_log.append({
                                'company': company_name,
                                'job_title': job_title,
                                'resume': resume_file,
                                'status': '✅ Sent'
                            })
                            
                        except Exception as e:
                            status_text.text(f"[{i+1}/{len(companies)}] ❌ Failed: {str(e)[:50]}")
                            failed_count += 1
                            sent_log.append({
                                'company': company_name,
                                'job_title': job_title,
                                'resume': resume_file,
                                'status': f'❌ {str(e)[:30]}'
                            })
                        
                        time.sleep(1)
                    
                    # Show results
                    st.success(f"✅ Complete! Sent: {sent_count}, Failed: {failed_count}")
                    
                    # Display results table
                    st.subheader("📊 Results")
                    result_data = []
                    for log in sent_log:
                        result_data.append({
                            'Company': log['company'],
                            'Job Title': log['job_title'],
                            'Resume': log['resume'],
                            'Status': log['status']
                        })
                    st.table(result_data)
                    
                    # Save log to Excel
                    log_wb = openpyxl.Workbook()
                    log_ws = log_wb.active
                    log_ws.title = "Sent Emails"
                    
                    log_ws['A1'] = "Company"
                    log_ws['B1'] = "Job Title"
                    log_ws['C1'] = "Resume"
                    log_ws['D1'] = "Date/Time"
                    log_ws['E1'] = "Status"
                    
                    for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
                        log_ws[cell].font = Font(bold=True, color="FFFFFF")
                        log_ws[cell].fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    
                    for row, log in enumerate(sent_log, 2):
                        log_ws[f'A{row}'] = log['company']
                        log_ws[f'B{row}'] = log['job_title']
                        log_ws[f'C{row}'] = log['resume']
                        log_ws[f'D{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        log_ws[f'E{row}'] = log['status']
                    
                    log_wb.save("sent_log.xlsx")
                    
                    with open("sent_log.xlsx", "rb") as f:
                        st.download_button(
                            label="📥 Download Results Log",
                            data=f.read(),
                            file_name="sent_log.xlsx",
                            use_container_width=True
                        )
                    
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")

with col2:
    st.info("""
    **📋 Before Sending:**
    1. ✅ Fill all fields in Configuration tab
    2. ✅ Create your email in Email Template tab
    3. ✅ Upload Excel file with companies
    4. ✅ Upload all resume PDFs
    5. ✅ Click START SENDING
    """)
